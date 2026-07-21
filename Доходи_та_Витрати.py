import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Створюємо нову робочу книгу
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Доходи & Витрати"

# Встановлюємо ширину колон
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12

# Стилі
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
income_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
expense_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Заголовок
ws['A1'] = "ОБЛІК ДОХОДІВ ТА ВИТРАТ"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

# Таблиця доходів
row = 3
ws['A3'] = "ДОХОДИ"
ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A3'].fill = income_fill

headers = ["Категорія", "Сума (₴)", "% від Доходу", "% від Витрат", "Тренд"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Дані доходів
incomes = [
    ("Основна зарплата", 50000),
    ("Фріланс проекти", 8000),
    ("Бонуси/Премії", 5000),
]

row = 4
total_income = 0
for category, amount in incomes:
    ws.cell(row=row, column=1).value = category
    ws.cell(row=row, column=2).value = amount
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).value = f"=B{row}/SUM($B$4:$B$6)"
    ws.cell(row=row, column=3).number_format = '0.0%'
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    total_income += amount
    row += 1

# Сума доходів
ws['A7'] = "ВСЬОГО ДОХОДІВ"
ws['A7'].font = Font(bold=True)
ws['A7'].fill = total_fill
ws['B7'] = f"=SUM(B4:B6)"
ws['B7'].font = Font(bold=True)
ws['B7'].fill = total_fill
ws['B7'].number_format = '#,##0'
ws['B7'].border = border
ws['A7'].border = border

# Таблиця витрат
row = 9
ws['A9'] = "ВИТРАТИ"
ws['A9'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A9'].fill = expense_fill

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Дані витрат
expenses = [
    ("Житло (оренда)", 12000),
    ("Комунальні услуги", 3200),
    ("Харчування", 2500),
    ("Транспорт", 1500),
    ("Інтернет/Мобільний", 800),
    ("Розваги", 1800),
    ("Інші витрати", 500),
]

row = 10
for category, amount in expenses:
    ws.cell(row=row, column=1).value = category
    ws.cell(row=row, column=2).value = amount
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).value = f"=B{row}/SUM($B$4:$B$6)"
    ws.cell(row=row, column=3).number_format = '0.0%'
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = border
    row += 1

# Сума витрат
ws['A17'] = "ВСЬОГО ВИТРАТ"
ws['A17'].font = Font(bold=True)
ws['A17'].fill = total_fill
ws['B17'] = f"=SUM(B10:B16)"
ws['B17'].font = Font(bold=True)
ws['B17'].fill = total_fill
ws['B17'].number_format = '#,##0'
ws['B17'].border = border
ws['A17'].border = border

# Чистий дохід
row = 19
ws['A19'] = "ЧИСТИЙ ДОХІД"
ws['A19'].font = Font(bold=True, size=12, color="FFFFFF")
ws['A19'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws['B19'] = "=B7-B17"
ws['B19'].font = Font(bold=True, size=12, color="FFFFFF")
ws['B19'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws['B19'].number_format = '#,##0'
ws['B19'].border = border
ws['A19'].border = border

# Збереження файлу
wb.save('Дохід_та_Витрати.xlsx')
print("✅ Файл 'Дохід_та_Витрати.xlsx' створений успішно!")
